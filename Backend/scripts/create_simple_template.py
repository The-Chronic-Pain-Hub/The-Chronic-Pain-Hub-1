import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Simplified template - 3 columns only
def create_simple_template():
    """
    Create simplified Excel template with only 3 columns:
    - Question_ID
    - Question_Text  
    - Scale_Values
    """
    
    output_path = Path(__file__).parent.parent.parent / 'Documents' / 'Module5_Codebook_Simple.xlsx'
    
    # Example data for PHQ-9 (as user provided)
    phq_example = {
        'Question_ID': [
            'PHQ_1', 'PHQ_2', 'PHQ_3', 'PHQ_4', 'PHQ_5', 
            'PHQ_6', 'PHQ_7', 'PHQ_8', 'PHQ_9'
        ],
        'Question_Text': [
            'Little interest or pleasure in doing things',
            'Feeling down, depressed or hopeless',
            'Trouble falling asleep, staying asleep, or sleeping too much',
            'Feeling tired or having little energy',
            'Poor appetite or overeating',
            'Feeling bad about yourself-or that you are a failure or have let yourself or your family down',
            'Trouble concentrating on things, such as reading the newspaper or watching television',
            'Moving or speaking so slowly that other people could have noticed. Or, the opposite being so fidgety or restless that you have been moving around a lot more than usual',
            'Thoughts that you would be better off dead or of hurting yourself in some way'
        ],
        'Scale_Values': [
            '1=Not at all, 2=Several Days, 3=More than half the days, 4=Nearly every day',
            '', '', '', '', '', '', '', ''  # Empty for other rows
        ]
    }
    
    # All scales list (IDENTITY first for pre-screening)
    scales = [
        ('IDENTITY (PRE-SCREEN)', 'IDENTITY', 3),  # Identity/role selection
        ('PHQ-9', 'PHQ', 9),
        ('IEQ', 'IEQ', 12),
        ('EDS', 'EDS', 9),
        ('PSEQ', 'PSEQ', 10),
        ('PB-SF', 'PB', None),
        ('SSPQ', 'SSPQ', 5),
        ('SOAPP', 'SOAPP', None),
        ('AFAQ', 'AFAQ', None),
        ('PCS', 'PCS', 13),
        ('CPAQ', 'CPAQ', 8),
        ('PMAQ', 'PMAQ', 14),
        ('TSK', 'TSK', None),
        ('P-FIBS', 'PFIBS', 9),
        ('PROMIS', 'PROMIS', 8),
        ('PI', 'PI', None),
        ('PGIS', 'PGIS', None),
        ('MAAS', 'MAAS', None),
        ('IPAQ', 'IPAQ', None),
        ('SSS', 'SSS', 8),
        ('PRS', 'PRS', None),
        ('FHQ', 'FHQ', None),
        ('MSPSS', 'MSPSS', None),
        ('SWLS', 'SWLS', None),
        ('PBS', 'PBS', None),
        ('WFC', 'WFC', None),
        ('CMNI', 'CMNI', None),
        ('PQ', 'PQ', None),
        ('FRSA', 'FRSA', None),
        ('MSE', 'MSE', None),
        ('ECON', 'ECON', 1),
        ('FSS', 'FSS', None),
        ('RAS', 'RAS', None),
        ('JSBR', 'JSBR', 5),
        ('ICS', 'ICS', None),
        ('HBI', 'HBI', 10),
    ]
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Instructions sheet
        instructions = {
            'Column': ['Question_ID', 'Question_Text', 'Scale_Values'],
            'Description': [
                'Question identifier from codebook (e.g., PHQ_1, IEQ_2, IDENTITY_1)',
                'Complete question text from codebook',
                'Scale values and labels (e.g., "1=Not at all, 2=Several Days"). Can fill in FIRST ROW (shared by all) OR EACH ROW (if questions have different options)'
            ],
            'Example': [
                'PHQ_1',
                'Little interest or pleasure in doing things',
                '1=Not at all, 2=Several Days, 3=More than half the days, 4=Nearly every day'
            ]
        }
        df_instructions = pd.DataFrame(instructions)
        df_instructions.to_excel(writer, sheet_name='README', index=False)
        
        # IDENTITY example (for pre-screening - with individual scales per question)
        identity_example = {
            'Question_ID': ['IDENTITY_1', 'IDENTITY_2', 'IDENTITY_3'],
            'Question_Text': [
                'What is your role?',
                'Have you experienced chronic pain (lasting 3+ months)?',
                'Are you currently receiving treatment for chronic pain?'
            ],
            'Scale_Values': [
                '1=Chronic Pain Patient, 2=Healthcare Provider, 3=Researcher, 4=Caregiver, 5=Other',
                '1=Yes currently, 2=Yes in the past, 3=No never',
                '1=Yes active treatment, 2=Yes but not active, 3=No'
            ]
        }
        df_identity = pd.DataFrame(identity_example)
        df_identity.to_excel(writer, sheet_name='IDENTITY (EXAMPLE)', index=False)
        
        # PHQ-9 example (filled - with shared scale)
        df_phq = pd.DataFrame(phq_example)
        df_phq.to_excel(writer, sheet_name='PHQ (EXAMPLE)', index=False)
        
        # Empty templates for other scales
        for scale_name, prefix, count in scales[2:]:  # Skip IDENTITY and PHQ since they're examples
            # Create empty rows based on expected count
            if count:
                empty_data = {
                    'Question_ID': [f'{prefix}_{i+1}' for i in range(count)],
                    'Question_Text': [''] * count,
                    'Scale_Values': [''] * count
                }
            else:
                # Unknown count - provide 15 empty rows
                empty_data = {
                    'Question_ID': [''] * 15,
                    'Question_Text': [''] * 15,
                    'Scale_Values': [''] * 15
                }
            
            df_scale = pd.DataFrame(empty_data)
            sheet_name = f"{prefix}"[:31]  # Excel sheet name limit
            df_scale.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Auto-adjust column widths
    workbook = load_workbook(output_path)
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width (add some padding)
            adjusted_width = min(max_length + 5, 100)  # Max 100 to avoid super wide columns
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_path)
    
    print("="*80)
    print("SIMPLIFIED EXCEL TEMPLATE CREATED")
    print("="*80)
    print(f"\nLocation: {output_path}")
    print(f"Total scales: {len(scales)}")
    print("\nFormat:")
    print("  - 3 columns only: Question_ID, Question_Text, Scale_Values")
    print("  - Column widths auto-adjusted")
    print("  - One sheet per scale")
    print("\nEXAMPLES:")
    print("  - IDENTITY (EXAMPLE): Each question has different options")
    print("  - PHQ (EXAMPLE): All questions share same scale (filled in row 1)")
    print("\nScale_Values填写规则:")
    print("  ✅ 如果某行Scale_Values是空的 → 使用第一行的scale（共享）")
    print("  ✅ 如果某行Scale_Values有内容 → 使用该行自己的scale（独立）")
    print("\nInstructions for researchers:")
    print("  1. Open the Excel file")
    print("  2. See 'README' sheet for column descriptions")
    print("  3. See examples for two different formats:")
    print("     - IDENTITY: Scale_Values filled in EACH row (different options)")
    print("     - PHQ: Scale_Values filled in FIRST row only (shared scale)")
    print("  4. For each scale sheet, fill from codebook:")
    print("     - Question_ID: Copy from codebook (e.g., IEQ_1, EDS_2)")
    print("     - Question_Text: Copy question text")
    print("     - Scale_Values:")
    print("       * Shared scale: Fill ONLY first row, other rows leave empty")
    print("       * Different options: Fill EACH row")
    print("  5. Delete unused rows if scale has fewer questions")
    print("  6. Save file")
    print("="*80)

if __name__ == '__main__':
    create_simple_template()
